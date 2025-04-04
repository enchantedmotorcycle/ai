import csv
import itertools
from langchain_core.output_parsers import StrOutputParser
from langchain_core.runnables import RunnablePassthrough
from langchain_core.prompts import PromptTemplate
from langchain_community.llms import Ollama
import argparse
import re
import datetime
from openpyxl import load_workbook, Workbook

parser = argparse.ArgumentParser(description='AI LLM parser.')
parser.add_argument('--question', type=str, help='Ask a question to LLM')
args = parser.parse_args()

llm = Ollama(model="mistral", base_url="http://192.168.2.176:11434")


# Function to load and filter CSV content based on vCPU and Memory
def load_and_filter_csv(file_path, vcpu, memory, os):
    rows = []
    with open(file_path, mode='r') as file:
        reader = csv.DictReader(itertools.islice(file, 5, None))  # Skip the first 5 lines
        #headers = reader.fieldnames
        #print(f"CSV Headers: {headers}")  # Print the headers to debug
        for row in reader:
            try:
                #print(row['License Model'])
                # if (int(row['vCPU']) >= vcpu and 
                #     float(row['Memory'].replace(' GiB', '')) >= memory and 
                #     int(row['vCPU']) <= (vcpu * 2) and 
                #     float(row['Memory'].replace(' GiB', '')) <= (memory * 2) and 
                if (vcpu <= int(row['vCPU']) <= (vcpu * 2) and 
                    memory <= float(row['Memory'].replace(' GiB', '')) <= (memory * 2) and 
                    row['TermType'] == "OnDemand" and 
                    row['Tenancy'] == "Shared" and 
                    row['Product Family'] == "Compute Instance" and 
                    #row['OfferingClass'] == "standard" and 
                    #row['PurchaseOption'] == "No Upfront" and 
                    row['License Model'] == "Bring your own license" and 
                    row['Operating System'] == os and 
                    row['Pre Installed S/W'] == "NA" and
                    row['Current Generation'] == "Yes" and
                    row['Instance Family'] == "Memory optimized" and
                    not row['usageType'].startswith("Reservation:")):
                    rows.append(row)
            except ValueError as e:
                #print(f"Skipping row due to error: {e}")
                continue
    
    # Sort rows by vCPU and Memory to pick the smallest instance
    rows.sort(key=lambda x: (int(x['vCPU']), float(x['Memory'].replace(' GiB', ''))), reverse=False)
    print(f"Filtered rows: {len(rows)}")
    return rows



# Define the output Excel file path
output_excel_path = f'EC2_Template.xlsx'

# Function to append data to an Excel file
def append_to_excel(file_path, data, sheet_name="Inputs", start_row=4):
    try:
        # Try to load the workbook if it exists
        workbook = load_workbook(file_path)
        # Access the specified sheet, or create it if it doesn't exist
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        # Write the header row
        sheet.append(list(data[0].keys())) # Convert dict keys to a list for the header

    # Append each row of data starting from the specified row
    for i, row in enumerate(data, start=start_row):
        for col_index, value in enumerate(row.values(), start=1):
            sheet.cell(row=i, column=col_index, value=value)

    # Save the workbook
    workbook.save(file_path)



# Function to load CSV content
def load_csv(file_path):
    with open(file_path, mode='r') as file:
        reader = csv.DictReader(file)
        rows = list(reader)
    return rows

# Function to format filtered CSV content and drop unwanted columns
def format_filtered_csv(rows):
    columns_to_keep = ['Instance Type', 'vCPU', 'Memory', 'PricePerUnit']
    formatted_rows = "\n".join([", ".join([f"{col}: {row[col]}" for col in columns_to_keep]) for row in rows])
    # # print the first 10 rows for debugging
    # for i, formatted_rows in enumerate(formatted_rows[:10]):
    #     print(f"Row {i+1}: {formatted_rows}")
    return formatted_rows

# Load AWS_EC2.csv content
aws_ec2_rows = load_csv('AWS_EC2.csv')

# Load request.csv content
request_rows = load_csv('request.csv')

# Add a new column to each row in the request_rows list
# for row in request_rows:
#     row['Raw JSON'] = 'None'
#     row['AWS Instance Type'] = 'None'
#     row['AWS Hourly Cost'] = 'None'
#     row['AWS Monthly Cost'] = 'None'

# Loop through each item in request.csv and filter AWS_EC2.csv content
for request in request_rows:
    atm_id = request['ATM ID']
    vcpu = int(request['vCPU'])
    memory = int(request['RAM GB'].replace(' GiB', ''))
    site = request['Site']
    env = request['ENV']
    zone = request['Zone']
    dbtype = request['DB Type']
    storage = request['Storage GB']
    if request['DB Type'] == "MSSQL":
        os = "Windows"
    else:
        os = "Linux"
    
    # Filter AWS_EC2.csv content based on the vCPU and Memory requirements
    filtered_csv_content = load_and_filter_csv('AWS_EC2.csv', vcpu, memory, os)
    
    # Format the filtered content
    formatted_filtered_csv_content = format_filtered_csv(filtered_csv_content)

    # Create the main prompt
    # main_prompt = PromptTemplate(
    #     input_variables=['question', 'context'],
    #     template="""
    #     You are an assistant for sizing AWS resources. Use the context to answer the question.
    #     Question: {question}
    #     Context: {context}
    #     Rules:
    #     1. Cost should be the most important factor.
    #     2. Pick the smallest instance that covers the vCPU and Memory requirements as close as possible.
    #     3. Do NOT select an instance that does not meet or exceed vCPU or Memory requirements in the question.
    #     4. Respond ONLY in a single line JSON format. ex: {{'Instance Type': 'xyz', 'vCPU': '4', 'Memory': '32', 'PricePerUnit': '0.34'}}
    #     Answer:
    #     """
    # )
    main_prompt = PromptTemplate(
        input_variables=['context'],
        template="""
        You are an assistant for sizing AWS resources. Use the context to answer the question.
        Context: {context}
        Rules:
        1. Cost should be the most important factor.
        2. Pick the smallest instance that covers the vCPU and Memory requirements as close as possible.
        3. Do NOT select an instance that does not meet or exceed vCPU or Memory requirements in the question.
        4. Respond ONLY in a single line JSON format. ex: {{'Instance Type': 'xyz', 'vCPU': '4', 'Memory': '32', 'PricePerUnit': '0.34'}}
        Answer:
        """
    )

    # rag_chain = (
    #     {"context": RunnablePassthrough(), "question": RunnablePassthrough()}
    #     | main_prompt
    #     | llm
    #     | StrOutputParser()
    # )
    rag_chain = (
        {"context": RunnablePassthrough()}
        | main_prompt
        | llm
        | StrOutputParser()
    )
    #print(formatted_filtered_csv_content)
    #result = rag_chain.invoke({"context": formatted_filtered_csv_content, "question": f"{args.question} vCPU: {vcpu} Memory: {memory} GiB"})
    input = f'"{formatted_filtered_csv_content} \n Question: \n {args.question} vCPU: {vcpu} Memory: {memory} GiB \n'
    result = rag_chain.invoke(input)
    print("Text response:")
    print(result)
    print("----------")
     # Extract JSON part from the response
    json_pattern = re.compile(r"\{.*?\}")
    match = json_pattern.search(result)
    print("JSON response:")
    if match:
        json_response = match.group(0)
        #print(type(json_response))
        print(json_response)
        #request['Raw JSON'] = json_response
        #request['AWS Instance Type'] = (eval(json_response))['Instance Type']
        #request['AWS Hourly Cost'] = (eval(json_response))['PricePerUnit']
        #request['AWS Monthly Cost'] = float((eval(json_response))['PricePerUnit']) * 730  # Assuming 730 hours in a month
        request['Group'] = atm_id
        request['Description'] = f"{site} - {env} - {zone} - {dbtype}"
        request['AWS Region'] = "us-east-1"
        if os == "Windows":
            request['Operating System'] = f"{os} Server"
        if os == "Linux":
            request['Operating System'] = f"{os}"
        request['Instance Type'] = (eval(json_response))['Instance Type']
        request['Tenancy'] = "Shared Instances"
        request['Number of Instances'] = "1"
        request['Assumed Usage'] = ""
        request['Usage Type'] = "Always On"
        request['Purchasing Option'] = "On-Demand"
        request['Storage Type'] = "General Purpose SSD (gp3)"
        request['Storage amount per Instance (GB)'] = f"{storage}"
        request['Provisioned IOPS per Instance'] = "50"
        request['EBS Throughput per Instance (applicable for gp3) (Mbps)'] = "500"
        request['Snapshot Frequency'] = "Daily"
        request['EBS Snapshot amount per Instance (GB/snapshot)'] = "0"
    else:
        print("No valid JSON response found.")
    print("----------")

# Generate a timestamp
timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')

# Add the timestamp to the output file name and write the modified request_rows to a new CSV file
output_file_path = f'modified_request_{timestamp}.csv'
with open(output_file_path, mode='w', newline='') as file:
    writer = csv.DictWriter(file, fieldnames=request_rows[0].keys())
    writer.writeheader()
    writer.writerows(request_rows)
print("Request rows")
print(request_rows)

# Remove specific columns (keys) from request_rows
columns_to_remove = ['Site', 'ENV', 'Zone', 'vCPU', 'RAM GB', 'DB Type', 'Storage GB']
request_rows = [{key: value for key, value in row.items() if key not in columns_to_remove} for row in request_rows]

#print(type(request_rows))
# Append the request_rows data to the Excel file
append_to_excel(output_excel_path, request_rows)